from __future__ import annotations

import argparse
import csv
import hashlib
import html as html_lib
import io
import json
import logging
import os
import re
import shutil
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from http.cookiejar import CookieJar
from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Sequence


PRICE_TOKEN_PATTERN = re.compile(
    r"(?P<value>\d+(?:[.,]\d+)?(?:[ \t]\d{3})*)[ \t]*(?P<currency>UZS|USD|RUB|EUR|so'?m|som|sum|сум|сўм|руб|₽|\$|€)",
    re.IGNORECASE,
)

LABELLED_PRICE_PATTERN = re.compile(
    r"(?:price|цена|narx|нарх)\s*[:\-]?\s*(?P<value>\d+(?:[.,]\d+)?(?:[ \t]\d{3})*)(?:[ \t]*(?P<currency>[A-Z]{3}|сум|so'?m|som|sum|\$|€|₽))?",
    re.IGNORECASE,
)

EXCEL_ILLEGAL_CHARACTERS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
SOURCE_SYSTEM = "milana_catalog_processor"
IMAGE_FILE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}

CODE_PATTERNS = [
    re.compile(r"\b(?P<code>[A-ZА-ЯΑ-ΩФΦ]{1,5}[-_/]?\d{2,6}(?:[-_/][A-ZА-ЯΑ-Ω0-9]{1,6})?)\b", re.IGNORECASE),
    re.compile(r"(?P<code>[\\/|I]-\d{2,6})", re.IGNORECASE),
    re.compile(
        r"(?:art(?:ikul)?|article|code|model|sku|код|артикул|модель|kodi|kod|modeli)\s*[:#\-]?\s*(?P<code>[A-ZА-Я0-9][A-ZА-Я0-9./_\-]{2,})",
        re.IGNORECASE,
    ),
    re.compile(r"\b(?P<code>[A-ZА-Я]{1,5}[-_/]?\d{2,6}(?:[-_/][A-ZА-Я0-9]{1,6})?)\b", re.IGNORECASE),
    re.compile(r"\b(?P<code>[A-Z0-9]{2,}[-_/][A-Z0-9][A-Z0-9\-_/]{2,})\b", re.IGNORECASE),
]


@dataclass
class TextBlock:
    bbox: tuple[int, int, int, int]
    text: str


@dataclass
class ProductRecord:
    catalog_date: str
    source_pdf: str
    source_pdf_path: str
    page: int
    card_index: int
    bbox: str
    model_code: str
    product_code: str
    material_type: str
    price: str
    currency: str
    is_visible: bool
    extraction_status: str
    native_text: str
    ocr_text: str
    combined_text: str
    image_path: str
    image_sha256: str
    image_fingerprint: str
    embedding_model: str
    embedding_path: str
    embedding_preview: str


@dataclass
class ProductFields:
    model_code: str
    product_code: str
    material_type: str
    price: str
    currency: str
    extraction_status: str


@dataclass
class CatalogSource:
    catalog_id: str
    title: str
    pdf_url: str


@dataclass
class DownloadedPdf:
    catalog_id: str
    title: str
    pdf_url: str
    file_id: str
    path: Path
    sha256: str
    changed: bool


class DependencyError(RuntimeError):
    pass


class OcrEngine:
    def __init__(self, lang: str, tesseract_cmd: str | None, logger: logging.Logger):
        self.lang = lang
        self.logger = logger
        self.available = False
        self.pytesseract = None
        self.rapid_ocr = None

        try:
            import pytesseract

            if tesseract_cmd:
                pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
            pytesseract.get_tesseract_version()
            self.pytesseract = pytesseract
        except Exception as exc:  # pragma: no cover - depends on local OCR install
            self.logger.warning("Tesseract is not available (%s). Trying RapidOCR fallback.", exc)

        try:
            from rapidocr_onnxruntime import RapidOCR

            self.rapid_ocr = RapidOCR()
        except Exception as exc:  # pragma: no cover - optional fallback dependency
            self.logger.warning("RapidOCR fallback is not available (%s).", exc)

        self.available = self.pytesseract is not None or self.rapid_ocr is not None

    def read(self, image) -> str:
        if not self.available:
            return ""

        text_parts: list[str] = []

        if self.pytesseract is not None:
            from PIL import Image, ImageOps

            scale = 2 if image.width < 1200 else 1
            work = image.convert("L")
            if scale > 1:
                resample = getattr(Image, "Resampling", Image).LANCZOS
                work = work.resize((work.width * scale, work.height * scale), resample)
            work = ImageOps.autocontrast(work)

            try:
                text_parts.append(self.pytesseract.image_to_string(work, lang=self.lang, config="--psm 6").strip())
            except Exception as exc:  # pragma: no cover - depends on local OCR languages
                if self.lang != "eng":
                    self.logger.warning("OCR failed with '%s' for one card; retrying with English OCR.", self.lang)
                    try:
                        text_parts.append(self.pytesseract.image_to_string(work, lang="eng", config="--psm 6").strip())
                    except Exception:
                        pass
                else:
                    self.logger.warning("OCR failed for one card: %s", exc)

        if self.rapid_ocr is not None:
            try:
                import numpy as np

                result, _ = self.rapid_ocr(np.array(image.convert("RGB")))
                if result:
                    text_parts.append("\n".join(item[1] for item in result if len(item) > 1 and item[1]))
            except Exception as exc:  # pragma: no cover - depends on OCR model runtime
                self.logger.warning("RapidOCR failed for one card: %s", exc)

        return combine_text(*text_parts)


class EmbeddingEngine:
    def __init__(self, enable_ml_embeddings: bool, logger: logging.Logger):
        self.logger = logger
        self.model = None
        self.model_name = "color-histogram"

        if not enable_ml_embeddings:
            return

        try:
            from sentence_transformers import SentenceTransformer

            self.model = SentenceTransformer("clip-ViT-B-32")
            self.model_name = "clip-ViT-B-32"
        except Exception as exc:  # pragma: no cover - optional heavy dependency
            self.logger.warning(
                "ML embeddings disabled: sentence-transformers/CLIP is not ready (%s). "
                "Using color-histogram vectors instead.",
                exc,
            )

    def vector_for(self, image) -> tuple[str, list[float]]:
        if self.model is not None:
            vector = self.model.encode([image.convert("RGB")], normalize_embeddings=True)[0]
            return self.model_name, [round(float(value), 8) for value in vector.tolist()]

        return self.model_name, color_histogram_embedding(image)


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(levelname)s: %(message)s",
    )
    logger = logging.getLogger("catalog_processor")

    try:
        result_path = process_catalogs(args, logger)
    except DependencyError as exc:
        logger.error("%s", exc)
        return 2
    except Exception:
        logger.exception("Catalog processing failed.")
        return 1

    logger.info("Done. Latest local artifact: %s", result_path)
    return 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Extract product cards from daily PDF catalogs and refresh Supabase or Excel outputs.",
    )
    parser.add_argument(
        "--source",
        choices=["auto", "website", "local"],
        default="website",
        help="PDF source. 'website' reads Google Drive links from index.html; 'local' reads --input.",
    )
    parser.add_argument("--website-index", default="index.html", help="HTML file containing the Google Drive PDF links.")
    parser.add_argument("--input", default="incoming_pdfs", help="Folder containing local PDF catalogs.")
    parser.add_argument("--output", default="outputs/catalog_processing", help="Folder for downloaded PDFs and local generated artifacts.")
    parser.add_argument(
        "--destination",
        choices=["supabase", "excel", "both"],
        default="supabase",
        help="Where to save refreshed product data.",
    )
    parser.add_argument("--workbook", default=None, help="Optional .xlsx output path.")
    parser.add_argument("--download-dir", default=None, help="Folder used to cache downloaded Google Drive PDFs.")
    parser.add_argument("--catalog-date", default=date.today().isoformat(), help="Date label saved to the workbook.")
    parser.add_argument("--expected-count", type=int, default=4, help="Warn if the PDF count is different.")
    parser.add_argument("--dpi", type=int, default=180, help="Render resolution used for extraction and OCR.")
    parser.add_argument(
        "--grid",
        default="auto",
        help="Card extraction mode. Use 'auto', '1x1', '2x2', '2x3', etc. for known card layouts.",
    )
    parser.add_argument("--ocr-lang", default="eng+rus+uzb", help="Tesseract OCR languages.")
    parser.add_argument("--tesseract-cmd", default=None, help="Full path to tesseract.exe if it is not on PATH.")
    parser.add_argument("--skip-ocr", action="store_true", help="Use only native PDF text extraction.")
    parser.add_argument(
        "--enable-ml-embeddings",
        action="store_true",
        help="Use CLIP image embeddings when optional ML dependencies are installed.",
    )
    parser.add_argument("--force", action="store_true", help="Reprocess PDFs even when downloaded files did not change.")
    parser.add_argument("--download-only", action="store_true", help="Download/check PDFs without rebuilding Excel.")
    parser.add_argument(
        "--locked-catalogs",
        default=None,
        help="Optional JSON file with source_pdfs to preserve instead of regenerating.",
    )
    parser.add_argument("--supabase-url", default=os.getenv("SUPABASE_URL"), help="Supabase project URL.")
    parser.add_argument(
        "--supabase-key",
        default=os.getenv("SUPABASE_SERVICE_ROLE_KEY") or os.getenv("SUPABASE_SECRET_KEY") or os.getenv("SUPABASE_KEY"),
        help="Supabase secret/service-role key. Prefer setting SUPABASE_SERVICE_ROLE_KEY in the environment.",
    )
    parser.add_argument(
        "--supabase-table",
        default=os.getenv("SUPABASE_PRODUCTS_TABLE", "milana_products"),
        help="Supabase table refreshed by the daily job.",
    )
    parser.add_argument(
        "--supabase-overrides-table",
        default=os.getenv("SUPABASE_OVERRIDES_TABLE", "milana_product_overrides"),
        help="Supabase table containing admin manual product overrides.",
    )
    parser.add_argument(
        "--supabase-image-bucket",
        default=os.getenv("SUPABASE_IMAGE_BUCKET", "product-images"),
        help="Supabase Storage bucket for product thumbnails.",
    )
    parser.add_argument(
        "--supabase-image-prefix",
        default=os.getenv("SUPABASE_IMAGE_PREFIX", "milana/latest"),
        help="Supabase Storage folder/prefix for latest product thumbnails.",
    )
    parser.add_argument("--verbose", action="store_true", help="Show debug logging.")
    return parser


def process_catalogs(args: argparse.Namespace, logger: logging.Logger) -> Path:
    check_required_dependencies()

    input_dir = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()
    workbook_path = (
        Path(args.workbook).resolve()
        if args.workbook
        else output_dir / "milana_products_latest.xlsx"
    )
    images_dir = output_dir / "images" / "latest"
    embeddings_dir = output_dir / "embeddings" / "latest"
    output_dir.mkdir(parents=True, exist_ok=True)
    images_dir.mkdir(parents=True, exist_ok=True)
    embeddings_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "milana_products_latest.json"
    locked_sources = load_locked_catalog_sources(args, output_dir)
    locked_payloads = load_locked_product_payloads(json_path, locked_sources)
    locked_image_backups = backup_locked_product_images(locked_payloads, output_dir)

    pdf_paths: list[Path]
    manifest_path = output_dir / "download_manifest.json"
    manifest_payload: dict | None = None
    source_changed = True

    if args.source in {"website", "auto"}:
        sources = read_catalog_sources(Path(args.website_index).resolve())
        if sources:
            download_dir = Path(args.download_dir).resolve() if args.download_dir else output_dir / "downloaded_pdfs" / "latest"
            previous_manifest = load_json(manifest_path)
            downloaded = download_catalog_sources(sources, download_dir, previous_manifest, logger)
            pdf_paths = [item.path for item in downloaded]
            source_changed = any(item.changed for item in downloaded) or not workbook_path.exists()
            manifest_payload = build_download_manifest(downloaded)
        elif args.source == "website":
            raise RuntimeError(f"No PDF links were found in {Path(args.website_index).resolve()}.")
        else:
            pdf_paths = sorted(input_dir.glob("*.pdf"))
    else:
        pdf_paths = sorted(input_dir.glob("*.pdf"))

    if args.expected_count and len(pdf_paths) != args.expected_count:
        logger.warning("Expected %s PDFs, found %s.", args.expected_count, len(pdf_paths))
    if not pdf_paths:
        logger.warning("No PDF files found.")

    if args.download_only:
        if manifest_payload is not None:
            save_json(manifest_path, manifest_payload)
        logger.info("Download/check complete. Outputs were not refreshed because --download-only was used.")
        return output_dir

    if not args.force and not source_changed:
        logger.info("No PDF changes detected. Output is already current.")
        return workbook_path if workbook_path.exists() else output_dir

    ocr_engine = None if args.skip_ocr else OcrEngine(args.ocr_lang, args.tesseract_cmd, logger)
    embedding_engine = EmbeddingEngine(args.enable_ml_embeddings, logger)

    reset_generated_outputs(output_dir, workbook_path, images_dir, embeddings_dir)
    restore_locked_product_images(locked_image_backups)
    external_image_replacements = load_external_image_replacements(output_dir, logger)

    records: list[ProductRecord] = []
    for pdf_path in pdf_paths:
        if pdf_path.name in locked_sources:
            logger.info("Skipping locked catalog %s", pdf_path.name)
            continue

        logger.info("Processing %s", pdf_path.name)
        records.extend(
            process_pdf(
                pdf_path=pdf_path,
                args=args,
                images_dir=images_dir,
                embeddings_dir=embeddings_dir,
                ocr_engine=ocr_engine,
                embedding_engine=embedding_engine,
                external_image_replacements=external_image_replacements,
                logger=logger,
            )
        )

    csv_path = output_dir / "milana_products_latest.csv"
    json_payload = sorted(
        [asdict(record) for record in records] + locked_payloads,
        key=lambda item: (
            str(item.get("source_pdf") or ""),
            int(item.get("page") or 0),
            int(item.get("card_index") or 0),
        ),
    )
    json_path.write_text(json.dumps(json_payload, indent=2, ensure_ascii=False), encoding="utf-8")

    if args.destination in {"supabase", "both"}:
        sync_records_to_supabase(records, args, output_dir, logger)

    if args.destination in {"excel", "both"}:
        write_excel(records, workbook_path)
        write_csv(records, csv_path)

    if manifest_payload is not None:
        save_json(manifest_path, manifest_payload)

    logger.info("Extracted %s product card candidates.", len(records))
    if locked_payloads:
        logger.info("Preserved %s locked product row(s).", len(locked_payloads))
    if args.destination in {"excel", "both"}:
        logger.info("Excel workbook: %s", workbook_path)
        logger.info("CSV copy: %s", csv_path)
    logger.info("JSON copy: %s", json_path)
    return workbook_path if args.destination in {"excel", "both"} else output_dir


def read_catalog_sources(index_path: Path) -> list[CatalogSource]:
    if not index_path.exists():
        return []

    text = index_path.read_text(encoding="utf-8", errors="replace")
    object_pattern = re.compile(r"\{id:(?P<id>\d+),(?P<body>.*?)\}", re.DOTALL)
    title_pattern = re.compile(r'title_en\s*:\s*"(?P<title>[^"]+)"')
    pdf_pattern = re.compile(r'pdf\s*:\s*"(?P<pdf>[^"]+)"')

    sources: list[CatalogSource] = []
    for match in object_pattern.finditer(text):
        body = match.group("body")
        pdf_match = pdf_pattern.search(body)
        if not pdf_match:
            continue
        title_match = title_pattern.search(body)
        title = title_match.group("title") if title_match else f"catalog_{match.group('id')}"
        sources.append(
            CatalogSource(
                catalog_id=match.group("id"),
                title=title,
                pdf_url=pdf_match.group("pdf"),
            )
        )

    if sources:
        return sources

    urls = pdf_pattern.findall(text)
    return [
        CatalogSource(catalog_id=str(index), title=f"catalog_{index}", pdf_url=url)
        for index, url in enumerate(urls, start=1)
    ]


def download_catalog_sources(
    sources: list[CatalogSource],
    download_dir: Path,
    previous_manifest: dict,
    logger: logging.Logger,
) -> list[DownloadedPdf]:
    download_dir.mkdir(parents=True, exist_ok=True)
    previous_hashes = {
        str(item.get("file_id") or item.get("pdf_url")): item.get("sha256")
        for item in previous_manifest.get("files", [])
    }
    downloaded: list[DownloadedPdf] = []

    for source in sources:
        file_id = google_drive_file_id(source.pdf_url) or source.pdf_url
        filename = f"{int(source.catalog_id):02d}_{safe_filename(source.title)}.pdf"
        target_path = download_dir / filename
        temp_path = target_path.with_suffix(".download")

        logger.info("Checking Google Drive catalog %s: %s", source.catalog_id, source.title)
        download_pdf(source.pdf_url, temp_path)
        if not is_pdf_file(temp_path):
            sample = temp_path.read_bytes()[:300].decode("utf-8", errors="replace")
            temp_path.unlink(missing_ok=True)
            raise RuntimeError(
                "Downloaded content is not a PDF. Check that the Google Drive file is public/downloadable. "
                f"Catalog {source.catalog_id}; first bytes: {sample!r}"
            )

        sha256 = sha256_file(temp_path)
        previous_sha256 = previous_hashes.get(file_id)
        changed = previous_sha256 != sha256
        shutil.move(str(temp_path), str(target_path))

        if changed:
            logger.info("Catalog %s changed or is new.", source.catalog_id)
        else:
            logger.info("Catalog %s is unchanged.", source.catalog_id)

        downloaded.append(
            DownloadedPdf(
                catalog_id=source.catalog_id,
                title=source.title,
                pdf_url=source.pdf_url,
                file_id=file_id,
                path=target_path,
                sha256=sha256,
                changed=changed,
            )
        )

    return downloaded


def download_pdf(url: str, target_path: Path) -> None:
    file_id = google_drive_file_id(url)
    download_url = google_drive_download_url(url, file_id)
    cookie_jar = CookieJar()
    opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cookie_jar))

    response, first_chunk = open_download_response(opener, download_url)
    if looks_like_pdf_chunk(first_chunk):
        save_stream(response, target_path, first_chunk)
        return

    preview = first_chunk + response.read(1024 * 1024)
    confirmation_url = google_drive_confirmation_url(preview.decode("utf-8", errors="ignore"), cookie_jar, file_id)
    if not confirmation_url:
        target_path.write_bytes(preview)
        return

    response, first_chunk = open_download_response(opener, confirmation_url)
    save_stream(response, target_path, first_chunk)


def open_download_response(opener, url: str):
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 MilanaCatalogProcessor/1.0",
            "Accept": "application/pdf,text/html,*/*",
        },
    )
    response = opener.open(request, timeout=120)
    first_chunk = response.read(32768)
    return response, first_chunk


def save_stream(response, target_path: Path, first_chunk: bytes) -> None:
    target_path.parent.mkdir(parents=True, exist_ok=True)
    with target_path.open("wb") as handle:
        handle.write(first_chunk)
        shutil.copyfileobj(response, handle)


def google_drive_file_id(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    query = urllib.parse.parse_qs(parsed.query)
    if query.get("id"):
        return query["id"][0]

    match = re.search(r"/d/([A-Za-z0-9_-]+)", parsed.path)
    if match:
        return match.group(1)

    match = re.search(r"[?&]id=([A-Za-z0-9_-]+)", url)
    return match.group(1) if match else ""


def google_drive_download_url(url: str, file_id: str, extra_params: dict[str, str] | None = None) -> str:
    if not file_id or "drive.google.com" not in url:
        if not file_id:
            return url
        base_url = "https://drive.google.com/uc"
    else:
        base_url = "https://drive.google.com/uc"

    query = {"export": "download", "id": file_id}
    if extra_params:
        query.update(extra_params)
    return base_url + "?" + urllib.parse.urlencode(query)


def google_drive_confirmation_url(html: str, cookie_jar: CookieJar, file_id: str) -> str:
    for cookie in cookie_jar:
        if cookie.name.startswith("download_warning"):
            return google_drive_download_url("", file_id, {"confirm": cookie.value})

    form_url = google_drive_form_download_url(html)
    if form_url:
        return form_url

    patterns = [
        r"confirm=([0-9A-Za-z_]+)",
        r'name="confirm"\s+value="([^"]+)"',
        r'"confirm"\s*,\s*"([^"]+)"',
    ]
    for pattern in patterns:
        match = re.search(pattern, html)
        if match:
            return google_drive_download_url("", file_id, {"confirm": match.group(1)})
    return ""


def google_drive_form_download_url(html: str) -> str:
    form_match = re.search(r'<form[^>]+id="download-form"[^>]+action="([^"]+)"[^>]*>(.*?)</form>', html, re.DOTALL)
    if not form_match:
        form_match = re.search(r'<form[^>]+action="([^"]+)"[^>]*>(.*?)</form>', html, re.DOTALL)
    if not form_match:
        return ""

    action = html_lib.unescape(form_match.group(1))
    body = form_match.group(2)
    params = {}
    for name, value in re.findall(r'<input[^>]+name="([^"]+)"[^>]+value="([^"]*)"', body):
        params[html_lib.unescape(name)] = html_lib.unescape(value)
    if not params:
        return ""
    return action + "?" + urllib.parse.urlencode(params)


def looks_like_pdf_chunk(chunk: bytes) -> bool:
    return chunk.lstrip().startswith(b"%PDF")


def is_pdf_file(path: Path) -> bool:
    with path.open("rb") as handle:
        return looks_like_pdf_chunk(handle.read(64))


def build_download_manifest(downloaded: list[DownloadedPdf]) -> dict:
    return {
        "updated_at": datetime.now().isoformat(timespec="seconds"),
        "files": [
            {
                "catalog_id": item.catalog_id,
                "title": item.title,
                "pdf_url": item.pdf_url,
                "file_id": item.file_id,
                "path": str(item.path),
                "sha256": item.sha256,
            }
            for item in downloaded
        ],
    }


def load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def save_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")


def load_locked_catalog_sources(args: argparse.Namespace, output_dir: Path) -> set[str]:
    path = Path(args.locked_catalogs).resolve() if args.locked_catalogs else output_dir / "locked_catalogs.json"
    if not path.exists():
        return set()

    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return set()

    values = payload.get("source_pdfs") if isinstance(payload, dict) else payload
    if not isinstance(values, list):
        return set()

    return {str(value) for value in values if str(value).strip()}


def load_locked_product_payloads(json_path: Path, locked_sources: set[str]) -> list[dict]:
    if not locked_sources or not json_path.exists():
        return []

    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return []

    if not isinstance(payload, list):
        return []

    return [
        item
        for item in payload
        if isinstance(item, dict) and item.get("source_pdf") in locked_sources
    ]


def backup_locked_product_images(locked_payloads: list[dict], output_dir: Path) -> list[tuple[Path, Path]]:
    backups: list[tuple[Path, Path]] = []
    if not locked_payloads:
        return backups

    backup_dir = output_dir / "_locked_image_backup"
    if backup_dir.exists():
        shutil.rmtree(backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)

    seen: set[Path] = set()
    for item in locked_payloads:
        for key in ("image_path", "embedding_path"):
            raw_path = str(item.get(key) or "")
            if not raw_path:
                continue

            path = Path(raw_path)
            if not path.exists() or path in seen:
                continue

            seen.add(path)
            backup_path = backup_dir / hashlib.sha256(str(path).encode("utf-8")).hexdigest()
            shutil.copy2(path, backup_path)
            backups.append((backup_path, path))

    return backups


def restore_locked_product_images(backups: list[tuple[Path, Path]]) -> None:
    for backup_path, original_path in backups:
        if not backup_path.exists():
            continue

        original_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(backup_path, original_path)


def reset_output_dir(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def reset_generated_outputs(output_dir: Path, workbook_path: Path, images_dir: Path, embeddings_dir: Path) -> None:
    reset_output_dir(output_dir / "images")
    reset_output_dir(output_dir / "embeddings")
    reset_output_dir(output_dir / "thumbnails")
    reset_output_dir(output_dir / "storage_images")
    images_dir.mkdir(parents=True, exist_ok=True)
    embeddings_dir.mkdir(parents=True, exist_ok=True)

    for path in [
        workbook_path,
        workbook_path.with_suffix(".csv"),
        workbook_path.with_suffix(".json"),
    ]:
        path.unlink(missing_ok=True)


def check_required_dependencies() -> None:
    missing: list[str] = []
    for module_name, package_name in [
        ("fitz", "PyMuPDF"),
        ("PIL", "Pillow"),
        ("openpyxl", "openpyxl"),
    ]:
        try:
            __import__(module_name)
        except Exception:
            missing.append(package_name)

    if missing:
        joined = ", ".join(missing)
        raise DependencyError(f"Missing Python packages: {joined}. Run: python -m pip install -r requirements.txt")


def process_pdf(
    pdf_path: Path,
    args: argparse.Namespace,
    images_dir: Path,
    embeddings_dir: Path,
    ocr_engine: OcrEngine | None,
    embedding_engine: EmbeddingEngine,
    external_image_replacements: dict[str, Path],
    logger: logging.Logger,
) -> list[ProductRecord]:
    import fitz

    doc = fitz.open(pdf_path)
    records: list[ProductRecord] = []

    try:
        for page_index, page in enumerate(doc, start=1):
            page_image = render_page(page, args.dpi)
            text_blocks = extract_text_blocks(page, args.dpi)
            regions = price_position_card_regions(text_blocks, page_image.size)
            if not regions:
                regions = detect_card_regions(page_image, args.grid, logger)
                regions = split_multi_product_regions(regions, page_image.size, text_blocks)
            logger.debug("Page %s: %s card candidate(s).", page_index, len(regions))

            for card_index, bbox in enumerate(regions, start=1):
                card_image = page_image.crop(bbox)
                native_text = text_for_region(text_blocks, bbox)
                native_fields = extract_product_fields(native_text)
                ocr_text = (
                    ocr_engine.read(card_image)
                    if ocr_engine and product_fields_need_ocr(native_fields)
                    else ""
                )
                combined_text = combine_text(ocr_text, native_text) if ocr_text else native_text
                fields = extract_product_fields(combined_text)
                if not fields.price or not fields.product_code:
                    continue

                image_stem = f"{safe_stem(pdf_path)}_p{page_index:03d}_c{card_index:03d}"
                replacement_image_path = external_image_replacements.get(image_stem)
                image_suffix = replacement_image_path.suffix.lower() if replacement_image_path else ".png"
                if image_suffix not in {".jpg", ".jpeg", ".png", ".webp"}:
                    image_suffix = ".jpg"
                image_name = f"{image_stem}{image_suffix}"
                image_path = images_dir / image_name
                output_image = save_product_image(card_image, image_path, replacement_image_path, pdf_path.name)
                image_sha256 = sha256_file(image_path)
                image_fingerprint = image_fingerprint_hash(output_image)

                embedding_model, embedding_vector = embedding_engine.vector_for(output_image)
                embedding_path = embeddings_dir / f"{image_path.stem}.json"
                embedding_payload = {
                    "source_pdf": pdf_path.name,
                    "page": page_index,
                    "card_index": card_index,
                    "model": embedding_model,
                    "vector": embedding_vector,
                }
                embedding_path.write_text(json.dumps(embedding_payload), encoding="utf-8")

                records.append(
                    ProductRecord(
                        catalog_date=args.catalog_date,
                        source_pdf=pdf_path.name,
                        source_pdf_path=str(pdf_path),
                        page=page_index,
                        card_index=card_index,
                        bbox=",".join(str(value) for value in bbox),
                        model_code=fields.model_code,
                        product_code=fields.product_code,
                        material_type=fields.material_type,
                        price=fields.price,
                        currency=fields.currency,
                        is_visible=True,
                        extraction_status=fields.extraction_status,
                        native_text=native_text,
                        ocr_text=ocr_text,
                        combined_text=combined_text,
                        image_path=str(image_path),
                        image_sha256=image_sha256,
                        image_fingerprint=image_fingerprint,
                        embedding_model=embedding_model,
                        embedding_path=str(embedding_path),
                        embedding_preview=json.dumps(embedding_vector[:12]),
                    )
                )
    finally:
        doc.close()

    return records


def save_product_image(card_image, image_path: Path, replacement_image_path: Path | None, source_pdf: str = ""):
    from PIL import Image, ImageOps

    if replacement_image_path and replacement_image_path.exists():
        with Image.open(replacement_image_path) as replacement:
            output_image = ImageOps.exif_transpose(replacement).convert("RGB")
    else:
        output_image = card_image.convert("RGB")
        output_image = remove_embedded_price_strip(output_image, source_pdf)

    suffix = image_path.suffix.lower()
    if suffix in {".jpg", ".jpeg"}:
        output_image.save(image_path, "JPEG", quality=92, optimize=True)
    elif suffix == ".png":
        output_image.save(image_path, "PNG", optimize=True)
    elif suffix == ".webp":
        output_image.save(image_path, "WEBP", quality=92, method=6)
    else:
        output_image.save(image_path)

    return output_image


def remove_embedded_price_strip(image, source_pdf: str):
    if source_pdf != "02_Milana_Man_Premium_Collection.pdf":
        return image

    width, height = image.size
    crop_y = int(height * 0.935)
    if height - crop_y < 24:
        return image
    return image.crop((0, 0, width, crop_y))


def load_external_image_replacements(output_dir: Path, logger: logging.Logger) -> dict[str, Path]:
    external_root = output_dir / "external_images"
    if not external_root.exists():
        return {}

    replacements: dict[str, Path] = {}
    skipped = 0

    for report_path in sorted(external_root.glob("*/website_replacement_report.csv")):
        source_lookup = {
            path.name.lower(): path
            for path in report_path.parent.rglob("*")
            if path.is_file() and path.suffix.lower() in IMAGE_FILE_SUFFIXES
        }

        with report_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                target_name = (row.get("website_image") or row.get("target_image") or "").strip()
                source_name = (row.get("source_image") or row.get("source_path") or "").strip()
                if not target_name or not source_name:
                    skipped += 1
                    continue

                source_path = Path(source_name)
                if not source_path.is_absolute():
                    source_path = source_lookup.get(source_path.name.lower(), Path())
                if not source_path.exists():
                    skipped += 1
                    continue

                replacements[Path(target_name).stem] = source_path

    if replacements:
        logger.info("Loaded %s external product image replacement(s).", len(replacements))
    if skipped:
        logger.debug("Skipped %s external image replacement row(s) without a source image.", skipped)
    return replacements


def render_page(page, dpi: int):
    import fitz
    from PIL import Image

    zoom = dpi / 72
    pixmap = page.get_pixmap(matrix=fitz.Matrix(zoom, zoom), alpha=False)
    return Image.open(io.BytesIO(pixmap.tobytes("png"))).convert("RGB")


def extract_text_blocks(page, dpi: int) -> list[TextBlock]:
    scale = dpi / 72
    blocks: list[TextBlock] = []

    text_dict = page.get_text("dict")
    for block in text_dict.get("blocks", []):
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            text = normalize_text(" ".join(str(span.get("text", "")) for span in spans))
            if not text:
                continue
            x0, y0, x1, y1 = line.get("bbox", (0, 0, 0, 0))
            blocks.append(TextBlock((int(x0 * scale), int(y0 * scale), int(x1 * scale), int(y1 * scale)), text))
    return blocks


def detect_card_regions(image, grid: str, logger: logging.Logger) -> list[tuple[int, int, int, int]]:
    width, height = image.size
    parsed_grid = parse_grid(grid)
    if parsed_grid:
        rows, cols = parsed_grid
        return [
            (
                int(col * width / cols),
                int(row * height / rows),
                int((col + 1) * width / cols),
                int((row + 1) * height / rows),
            )
            for row in range(rows)
            for col in range(cols)
        ]

    try:
        return auto_detect_card_regions(image, logger)
    except Exception as exc:  # pragma: no cover - optional OpenCV path
        logger.debug("Auto card detection failed, using full-page cards: %s", exc)
        return [(0, 0, width, height)]


def price_position_card_regions(
    text_blocks: list[TextBlock],
    image_size: tuple[int, int],
) -> list[tuple[int, int, int, int]]:
    width, height = image_size
    price_items: list[tuple[float, float, int]] = []
    for block in text_blocks:
        if not PRICE_TOKEN_PATTERN.search(block.text):
            continue
        x0, y0, x1, y1 = block.bbox
        price_items.append(((x0 + x1) / 2, (y0 + y1) / 2, y1))

    if len(price_items) < 2:
        return []

    row_centers = cluster_axis([item[1] for item in price_items], max(40, height * 0.10))
    col_centers = cluster_axis([item[0] for item in price_items], max(40, width * 0.12))
    if not row_centers or not col_centers:
        return []
    if len(row_centers) * len(col_centers) > len(price_items) + max(2, len(price_items) // 2):
        return []

    x_bounds = [index * width / len(col_centers) for index in range(len(col_centers) + 1)]
    row_price_bottoms = [0 for _ in row_centers]
    for _, y_center, y1 in price_items:
        row = nearest_axis_index(row_centers, y_center)
        row_price_bottoms[row] = max(row_price_bottoms[row], y1)

    y_bounds: list[float] = [0]
    row_gap = max(20, height * 0.012)
    for row in range(len(row_centers) - 1):
        y_bounds.append(min(height, row_price_bottoms[row] + row_gap))
    y_bounds.append(height)

    occupied: set[tuple[int, int]] = set()
    for x_center, y_center, _ in price_items:
        col = nearest_axis_index(col_centers, x_center)
        row = nearest_axis_index(row_centers, y_center)
        occupied.add((row, col))

    return [
        (
            int(x_bounds[col]),
            int(y_bounds[row]),
            int(x_bounds[col + 1]),
            int(y_bounds[row + 1]),
        )
        for row, col in sorted(occupied)
    ]


def cluster_axis(values: list[float], threshold: float) -> list[float]:
    clusters: list[list[float]] = []
    for value in sorted(values):
        if not clusters or abs(value - (sum(clusters[-1]) / len(clusters[-1]))) > threshold:
            clusters.append([value])
        else:
            clusters[-1].append(value)
    return [sum(cluster) / len(cluster) for cluster in clusters]


def axis_bounds(centers: list[float], start: int, end: int) -> list[float]:
    centers = sorted(centers)
    bounds: list[float] = [start]
    for previous, current in zip(centers, centers[1:]):
        bounds.append((previous + current) / 2)
    bounds.append(end)
    return bounds


def nearest_axis_index(centers: list[float], value: float) -> int:
    return min(range(len(centers)), key=lambda index: abs(centers[index] - value))


def split_multi_product_regions(
    regions: list[tuple[int, int, int, int]],
    image_size: tuple[int, int],
    text_blocks: list[TextBlock],
) -> list[tuple[int, int, int, int]]:
    page_area = image_size[0] * image_size[1]
    output: list[tuple[int, int, int, int]] = []

    for region in regions:
        region_text = text_for_region(text_blocks, region)
        price_count = count_prices(region_text)
        if price_count <= 1 or box_area(region) < page_area * 0.35:
            output.append(region)
            continue

        rows, cols = grid_for_product_count(price_count, region)
        output.extend(split_region(region, rows, cols))

    return output


def count_prices(text: str) -> int:
    return len(PRICE_TOKEN_PATTERN.findall(text))


def grid_for_product_count(count: int, region: tuple[int, int, int, int]) -> tuple[int, int]:
    width = max(1, region[2] - region[0])
    height = max(1, region[3] - region[1])

    if count == 2:
        return (1, 2) if width >= height * 0.8 else (2, 1)
    if count in {3, 4}:
        return 2, 2
    if count in {5, 6}:
        return 2, 3
    if count in {7, 8}:
        return 2, 4

    cols = max(1, round(count ** 0.5))
    rows = max(1, (count + cols - 1) // cols)
    return rows, cols


def split_region(region: tuple[int, int, int, int], rows: int, cols: int) -> list[tuple[int, int, int, int]]:
    x0, y0, x1, y1 = region
    width = x1 - x0
    height = y1 - y0
    output = []
    for row in range(rows):
        for col in range(cols):
            output.append(
                (
                    int(x0 + col * width / cols),
                    int(y0 + row * height / rows),
                    int(x0 + (col + 1) * width / cols),
                    int(y0 + (row + 1) * height / rows),
                )
            )
    return output


def parse_grid(value: str) -> tuple[int, int] | None:
    if value.lower() == "auto":
        return None
    match = re.fullmatch(r"(\d+)x(\d+)", value.lower().strip())
    if not match:
        raise ValueError("--grid must be 'auto' or a value like '2x2'.")
    rows = int(match.group(1))
    cols = int(match.group(2))
    if rows <= 0 or cols <= 0:
        raise ValueError("--grid rows and columns must be positive.")
    return rows, cols


def auto_detect_card_regions(image, logger: logging.Logger) -> list[tuple[int, int, int, int]]:
    import cv2
    import numpy as np

    width, height = image.size
    array = np.array(image.convert("RGB"))
    gray = cv2.cvtColor(array, cv2.COLOR_RGB2GRAY)
    gray = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(gray, 40, 140)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (11, 11))
    closed = cv2.morphologyEx(edges, cv2.MORPH_CLOSE, kernel, iterations=2)
    contours, _ = cv2.findContours(closed, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    page_area = width * height
    candidates: list[tuple[int, int, int, int]] = []
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        area = w * h
        aspect = w / max(h, 1)
        if area < page_area * 0.015 or area > page_area * 0.92:
            continue
        if w < 120 or h < 120:
            continue
        if not 0.22 <= aspect <= 4.5:
            continue
        pad = 10
        candidates.append((max(0, x - pad), max(0, y - pad), min(width, x + w + pad), min(height, y + h + pad)))

    candidates = remove_nested_boxes(candidates)
    candidates = sorted(candidates, key=lambda box: (box[1], box[0]))
    if candidates:
        return candidates

    logger.debug("No card contours found; falling back to one card per page.")
    return [(0, 0, width, height)]


def remove_nested_boxes(boxes: Iterable[tuple[int, int, int, int]]) -> list[tuple[int, int, int, int]]:
    result: list[tuple[int, int, int, int]] = []
    sorted_boxes = sorted(boxes, key=box_area, reverse=True)
    for box in sorted_boxes:
        if any(overlap_ratio(box, kept) > 0.88 for kept in result):
            continue
        result.append(box)
    return result


def box_area(box: tuple[int, int, int, int]) -> int:
    return max(0, box[2] - box[0]) * max(0, box[3] - box[1])


def overlap_ratio(a: tuple[int, int, int, int], b: tuple[int, int, int, int]) -> float:
    x0 = max(a[0], b[0])
    y0 = max(a[1], b[1])
    x1 = min(a[2], b[2])
    y1 = min(a[3], b[3])
    inter = box_area((x0, y0, x1, y1))
    return inter / max(1, min(box_area(a), box_area(b)))


def text_for_region(blocks: list[TextBlock], region: tuple[int, int, int, int]) -> str:
    selected: list[tuple[int, int, str]] = []
    rx0, ry0, rx1, ry1 = region
    for block in blocks:
        x0, y0, x1, y1 = block.bbox
        cx = (x0 + x1) / 2
        cy = (y0 + y1) / 2
        if rx0 <= cx <= rx1 and ry0 <= cy <= ry1:
            selected.append((y0, x0, block.text))
    return "\n".join(text for _, _, text in sorted(selected))


def combine_text(*parts: str) -> str:
    seen: set[str] = set()
    output: list[str] = []
    for part in parts:
        for line in part.splitlines():
            line = normalize_text(line)
            key = line.lower()
            if line and key not in seen:
                output.append(line)
                seen.add(key)
    return "\n".join(output)


def extract_product_fields(text: str) -> ProductFields:
    material_type = extract_material_type(text)
    lines = meaningful_lines(text)
    flat_text = "\n".join(lines)
    price, currency, price_line_index = extract_price_with_position(lines, flat_text)
    labelled_model, labelled_product = extract_labelled_model_and_product(lines)

    model_code = ""
    product_code = ""
    if labelled_model or labelled_product:
        if labelled_product:
            model_code = labelled_model if labelled_model != labelled_product else ""
            product_code = labelled_product
        else:
            product_code = labelled_model
    else:
        code_candidates = code_candidates_before_price(lines, price_line_index)
        if code_candidates:
            model_code, product_code = choose_model_and_product_code(code_candidates)
        else:
            product_code = extract_product_code(flat_text)

    status_parts = []
    if not product_code:
        status_parts.append("missing_code")
    if not price:
        status_parts.append("missing_price")
    if product_code and product_code.isdigit():
        status_parts.append("prefix_needs_ocr")
    if model_code and model_code.isdigit():
        status_parts.append("model_prefix_needs_ocr")

    return ProductFields(
        model_code=model_code,
        product_code=product_code,
        material_type=material_type,
        price=price,
        currency=currency,
        extraction_status=";".join(status_parts) or "ok",
    )


def product_fields_need_ocr(fields: ProductFields) -> bool:
    if not fields.price or not fields.product_code:
        return True
    return "needs_ocr" in fields.extraction_status


def extract_labelled_model_and_product(lines: list[str]) -> tuple[str, str]:
    model_code = ""
    product_code = ""

    for index, line in enumerate(lines):
        if is_model_label(line) and not model_code:
            model_code = first_code_after_label(lines, index)
        if is_code_label(line) and not product_code:
            product_code = first_code_after_label(lines, index)

    if model_code and not product_code:
        product_code = first_related_product_code(lines, model_code)

    return model_code, product_code


def first_code_after_label(lines: list[str], label_index: int) -> str:
    for line in lines[label_index + 1 : label_index + 6]:
        if is_model_label(line) or is_code_label(line):
            continue
        candidates = remove_duplicate_numeric_suffix_codes(extract_codes_from_line(line))
        for candidate in candidates:
            if candidate and not is_size_line(candidate):
                return candidate
    return ""


def first_related_product_code(lines: list[str], model_code: str) -> str:
    model_suffix = numeric_suffix(model_code)
    for line in lines:
        for candidate in extract_codes_from_line(line):
            if candidate == model_code:
                continue
            if candidate.isdigit() and model_suffix and candidate == model_suffix:
                continue
            if candidate and not candidate.isdigit() and not is_model_like_code(candidate):
                return candidate
    return ""


def numeric_suffix(value: str) -> str:
    match = re.search(r"(\d{2,6})$", value)
    return match.group(1) if match else ""


def is_model_label(value: str) -> bool:
    label = normalize_label(value)
    return label in {"MODE", "M0DE"} or "MODEL" in label or "MODEI" in label or label.endswith("ODEL")


def is_code_label(value: str) -> bool:
    label = normalize_label(value)
    return "CODE" in label or "C0DE" in label or label in {"KOD", "KODI"}


def normalize_label(value: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", value.upper())


def choose_model_and_product_code(candidates: list[str]) -> tuple[str, str]:
    normalized = [
        normalize_code(candidate)
        for candidate in candidates
        if normalize_code(candidate) and not is_size_line(normalize_code(candidate))
    ]
    normalized = remove_duplicate_numeric_suffix_codes(normalized)
    if not normalized:
        return "", ""
    if len(normalized) == 1:
        return "", normalized[0]

    previous = normalized[-2]
    last = normalized[-1]
    if is_model_like_code(last) and previous.isdigit():
        return last, previous
    return previous, last


def remove_duplicate_numeric_suffix_codes(codes: list[str]) -> list[str]:
    prefixed_suffixes = {
        match.group(1)
        for code in codes
        if not code.isdigit()
        for match in [re.search(r"(\d{2,6})$", code)]
        if match
    }
    return [code for code in codes if not (code.isdigit() and code in prefixed_suffixes)]


def is_model_like_code(value: str) -> bool:
    return bool(re.match(r"^(PJ|PM|MP|MT|MN|MF|SJ|TJ|F)-?\d+", value, re.IGNORECASE))


def extract_price(text: str) -> tuple[str, str]:
    lines = meaningful_lines(text)
    flat_text = "\n".join(lines)
    price, currency, _ = extract_price_with_position(lines, flat_text)
    return price, currency


def extract_price_with_position(lines: list[str], flat_text: str) -> tuple[str, str, int | None]:
    for index, line in enumerate(lines):
        matches = list(PRICE_TOKEN_PATTERN.finditer(line))
        if matches:
            match = matches[-1]
            return clean_number(match.group("value")), normalize_currency(match.group("currency")), index

        labelled = LABELLED_PRICE_PATTERN.search(line)
        if labelled:
            return (
                clean_number(labelled.group("value")),
                normalize_currency(labelled.groupdict().get("currency") or ""),
                index,
            )

    matches = list(PRICE_TOKEN_PATTERN.finditer(flat_text))
    if matches:
        match = matches[-1]
        return clean_number(match.group("value")), normalize_currency(match.group("currency")), None

    labelled = LABELLED_PRICE_PATTERN.search(flat_text)
    if labelled:
        return clean_number(labelled.group("value")), normalize_currency(labelled.groupdict().get("currency") or ""), None

    return "", "", None


def code_candidates_before_price(lines: list[str], price_line_index: int | None) -> list[str]:
    if price_line_index is None:
        search_lines = lines
    else:
        search_lines = lines[:price_line_index]

    candidates: list[str] = []
    for line in search_lines:
        for candidate in extract_codes_from_line(line):
            if candidate not in candidates:
                candidates.append(candidate)
    return candidates


def extract_codes_from_line(line: str) -> list[str]:
    cleaned = remove_price_tokens(line)
    cleaned = cleaned.replace("Î¦", "F").replace("Ð¤", "F")
    candidates: list[str] = []
    for match in re.finditer(r"\bM[^A-Z0-9\s-]{1,6}-(18\d{3})\b", cleaned, re.IGNORECASE):
        code = normalize_code("MP-" + match.group(1))
        if code not in candidates:
            candidates.append(code)
    if re.fullmatch(r"\d{3,6}", cleaned.strip()):
        return [cleaned.strip()]

    for pattern in CODE_PATTERNS:
        for match in pattern.finditer(cleaned):
            code = normalize_code(match.group("code"))
            if is_plausible_code(code) and code not in candidates:
                candidates.append(code)
    return candidates


def extract_product_code(text: str) -> str:
    for line in meaningful_lines(text):
        candidates = extract_codes_from_line(line)
        if candidates:
            return candidates[-1]
    return ""


def meaningful_lines(text: str) -> list[str]:
    lines: list[str] = []
    for raw_line in text.splitlines():
        line = normalize_text(raw_line)
        if not line or is_material_line(line) or is_packaging_line(line) or is_size_line(line):
            continue
        lines.append(line)
    return lines


def extract_material_type(text: str) -> str:
    upper_text = material_search_text(text)
    if has_any(upper_text, ("LAPSHA", "ЛАПША", "NANWA", "ЛANWA")):
        return "Lapsha"
    if has_any(upper_text, ("INTERLOCK SUPREM", "INTERLOCK SUPREME", "ИНТЕРЛОК СУПРЕМ", "WHTEPNOK CYNPEM")):
        return "Interlock Suprem"
    if has_any(upper_text, ("MUSLIN", "МУСЛИН", "MYCNUH", "MYCNNH", "MYCNUN")):
        return "Muslin"
    if has_any(upper_text, ("WAFFLE", "WAFLE", "WAFFEL")):
        return "Waffle"
    if has_any(upper_text, ("INTERLOCK", "ИНТЕРЛОК", "WHTEPNOK", "NTERLOCK")):
        return "Interlock"
    if has_any(upper_text, ("BAMBUK", "BAMBOO", "БАМБУК")):
        return "Bambuk"
    if has_any(upper_text, ("VISCOSE", "VISCOSA", "ВИСКОЗ")):
        return "Viscose"
    if has_any(upper_text, ("POLYESTER", "POLIESTER", "ПОЛИЭСТЕР")):
        return "Polyester"
    if has_any(upper_text, ("RIBANA", "RIBAN", "РИБАН")):
        return "Ribana"
    if has_any(upper_text, ("FLEECE", "FLIS", "FLISLI", "ФЛИС")):
        return "Fleece"
    if has_any(upper_text, ("VELOUR", "VELUR", "ВЕЛЮР")):
        return "Velour"
    if has_any(upper_text, ("SUPREM", "SUPREME", "СУПРЕМ", "CYNPEM")):
        return "Suprem"
    return "Suprem"


def material_search_text(text: str) -> str:
    normalized = normalize_text(text).upper()
    return re.sub(r"[^A-ZА-ЯЁ0-9%]+", " ", normalized)


def has_any(text: str, aliases: tuple[str, ...]) -> bool:
    return any(alias in text for alias in aliases)


def remove_price_tokens(text: str) -> str:
    return PRICE_TOKEN_PATTERN.sub(" ", LABELLED_PRICE_PATTERN.sub(" ", text))


def normalize_code(value: str) -> str:
    code = value.replace("Î¦", "F").replace("Ð¤", "F").strip(" .,:;").upper()
    code = code.translate(
        str.maketrans(
            {
                "А": "A",
                "В": "B",
                "Е": "E",
                "К": "K",
                "М": "M",
                "Н": "H",
                "О": "O",
                "Р": "P",
                "С": "C",
                "Т": "T",
                "Х": "X",
                "У": "Y",
                "П": "P",
                "Ф": "F",
                "Φ": "F",
            }
        )
    )
    code = code.replace("–", "-").replace("—", "-").replace("−", "-")
    code = re.sub(r"\s+", "", code)
    code = re.sub(r"^[\\/|I]-", "V-", code)
    code = re.sub(r"^MN-(\d+)$", r"MP-\1", code)
    code = re.sub(r"^M-(18\d{3})$", r"MP-\1", code)
    return code


def is_plausible_code(value: str) -> bool:
    if not value:
        return False
    if is_size_line(value):
        return False
    if "%" in value or value.upper() in {"COTTON", "LYCRA", "SUPREM", "POLYESTER", "POLYESTR"}:
        return False
    return bool(re.search(r"\d", value))


def is_size_line(value: str) -> bool:
    cleaned = remove_price_tokens(value)
    digits = re.sub(r"\D", "", cleaned)
    if len(digits) not in {2, 4, 6}:
        return False
    if len(digits) != len(cleaned.strip()):
        return False
    sizes = [int(digits[index : index + 2]) for index in range(0, len(digits), 2)]
    return all(34 <= size <= 64 for size in sizes)


def is_material_line(value: str) -> bool:
    upper = value.upper()
    material_words = ("COTTON", "LYCRA", "SUPREM", "POLYESTER", "POLYESTR")
    return "%" in upper or any(word in upper for word in material_words)


def is_packaging_line(value: str) -> bool:
    upper = value.upper()
    return "ШТ" in upper or "БРИКЕТ" in upper


def clean_number(value: str) -> str:
    value = re.sub(r"\s+", "", value)
    value = value.replace(",", ".") if value.count(",") == 1 and value.count(".") == 0 else value
    return value.strip(".")


def normalize_currency(value: str) -> str:
    normalized = value.strip().lower()
    if normalized in {"$", "usd"}:
        return "USD"
    if normalized in {"€", "eur"}:
        return "EUR"
    if normalized in {"₽", "rub", "руб"}:
        return "RUB"
    if normalized in {"uzs", "som", "sum", "so'm", "som", "сум", "сўм"}:
        return "UZS"
    return value.upper()


def normalize_text(text: str) -> str:
    return re.sub(r"[ \t\r\f\v]+", " ", text.replace("\x00", " ")).strip()


def image_fingerprint_hash(image) -> str:
    try:
        import imagehash

        return f"phash:{imagehash.phash(image)}"
    except Exception:
        return f"ahash:{average_hash(image)}"


def average_hash(image, hash_size: int = 16) -> str:
    from PIL import Image

    resample = getattr(Image, "Resampling", Image).LANCZOS
    pixels = list(image.convert("L").resize((hash_size, hash_size), resample).getdata())
    average = sum(pixels) / len(pixels)
    bits = "".join("1" if pixel >= average else "0" for pixel in pixels)
    return f"{int(bits, 2):0{hash_size * hash_size // 4}x}"


def color_histogram_embedding(image, bins: int = 8) -> list[float]:
    try:
        import numpy as np

        array = np.array(image.convert("RGB").resize((256, 256))).reshape(-1, 3)
        hist, _ = np.histogramdd(array, bins=(bins, bins, bins), range=((0, 256), (0, 256), (0, 256)))
        vector = hist.flatten().astype("float64")
        norm = np.linalg.norm(vector)
        if norm:
            vector = vector / norm
        return [round(float(value), 8) for value in vector.tolist()]
    except Exception:
        histogram = image.convert("RGB").resize((64, 64)).histogram()
        total = sum(histogram) or 1
        return [round(value / total, 8) for value in histogram]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_stem(path: Path) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", path.stem).strip("._") or "catalog"


def safe_filename(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value).strip("._") or "catalog"


def sync_records_to_supabase(
    records: list[ProductRecord],
    args: argparse.Namespace,
    output_dir: Path,
    logger: logging.Logger,
) -> None:
    if not args.supabase_url or not args.supabase_key:
        raise RuntimeError(
            "Supabase output needs SUPABASE_URL and SUPABASE_SERVICE_ROLE_KEY "
            "environment variables, or --supabase-url and --supabase-key."
        )

    client = SupabaseRestClient(args.supabase_url, args.supabase_key)
    run_id = f"{args.catalog_date}-{datetime.now():%H%M%S}"
    storage_image_dir = output_dir / "storage_images" / "latest"
    storage_image_dir.mkdir(parents=True, exist_ok=True)
    overrides = load_supabase_overrides(client, args.supabase_overrides_table, logger)
    refreshed_sources = sorted({record.source_pdf for record in records})
    if not refreshed_sources:
        logger.info("No unlocked records to sync to Supabase.")
        return

    logger.info("Preparing %s Supabase image upload(s).", len(records))
    rows = []
    for index, record in enumerate(records, start=1):
        storage_image_path = create_supabase_image(Path(record.image_path), storage_image_dir)
        storage_path = "/".join(
            [
                normalize_storage_prefix(args.supabase_image_prefix),
                safe_filename(storage_image_path.name),
            ]
        )
        client.upload_storage_object(args.supabase_image_bucket, storage_path, storage_image_path, "image/jpeg")

        row = supabase_row(record, run_id, args.supabase_image_bucket, storage_path, client.public_storage_url)
        apply_supabase_override(row, overrides.get(override_key(record.source_pdf, record.page, record.card_index)))
        rows.append(row)
        if index == 1 or index % 25 == 0 or index == len(records):
            logger.info("Uploaded %s/%s product image(s) to Supabase Storage.", index, len(records))

    logger.info("Deleting old Supabase rows from '%s' for %s source PDF(s).", args.supabase_table, len(refreshed_sources))
    for source_pdf in refreshed_sources:
        client.delete_table_rows(
            args.supabase_table,
            f"source_system=eq.{SOURCE_SYSTEM}&source_pdf=eq.{urllib.parse.quote(source_pdf, safe='')}",
        )
    logger.info("Inserting %s fresh Supabase row(s).", len(rows))
    client.insert_rows(args.supabase_table, rows)
    logger.info("Supabase refresh complete: %s row(s), %s image(s).", len(rows), len(rows))


def load_supabase_overrides(
    client: "SupabaseRestClient",
    table: str,
    logger: logging.Logger,
) -> dict[tuple[str, int, int], dict]:
    try:
        payload = client.request(
            "GET",
            f"{client.rest_url}/{urllib.parse.quote(table)}?select=*",
        )
    except Exception as exc:
        logger.warning("Manual overrides were not loaded from '%s': %s", table, exc)
        return {}

    overrides = json.loads(payload.decode("utf-8")) if payload else []
    return {
        override_key(item.get("source_pdf"), item.get("page"), item.get("card_index")): item
        for item in overrides
        if item.get("source_pdf") and item.get("page") is not None and item.get("card_index") is not None
    }


def override_key(source_pdf: str | None, page: int | None, card_index: int | None) -> tuple[str, int, int]:
    return (source_pdf or "", int(page or 0), int(card_index or 0))


def apply_supabase_override(row: dict, override: dict | None) -> None:
    if not override:
        return

    for field in [
        "model_code",
        "product_code",
        "material_type",
        "price",
        "currency",
        "image_url",
        "image_storage_bucket",
        "image_storage_path",
        "is_visible",
    ]:
        if field == "is_visible":
            if override.get(field) is not None:
                row[field] = bool(override[field])
        elif field in override:
            row[field] = override[field]


class SupabaseRestClient:
    def __init__(self, url: str, key: str):
        self.url = url.rstrip("/")
        self.key = key

    @property
    def rest_url(self) -> str:
        return f"{self.url}/rest/v1"

    @property
    def storage_url(self) -> str:
        return f"{self.url}/storage/v1"

    @property
    def public_storage_url(self) -> str:
        return f"{self.storage_url}/object/public"

    def delete_table_rows(self, table: str, filter_query: str) -> None:
        self.request(
            "DELETE",
            f"{self.rest_url}/{urllib.parse.quote(table)}?{filter_query}",
            headers={"Prefer": "return=minimal"},
        )

    def insert_rows(self, table: str, rows: list[dict], batch_size: int = 250) -> None:
        for start in range(0, len(rows), batch_size):
            batch = rows[start : start + batch_size]
            self.request(
                "POST",
                f"{self.rest_url}/{urllib.parse.quote(table)}",
                payload=batch,
                headers={"Prefer": "return=minimal"},
            )

    def upload_storage_object(self, bucket: str, object_path: str, local_path: Path, content_type: str) -> None:
        data = local_path.read_bytes()
        for attempt in range(1, 4):
            try:
                self.request(
                    "POST",
                    f"{self.storage_url}/object/{urllib.parse.quote(bucket)}/{quote_storage_path(object_path)}",
                    body=data,
                    headers={
                        "Content-Type": content_type,
                        "x-upsert": "true",
                    },
                )
                return
            except RuntimeError:
                if attempt >= 3:
                    raise
                time.sleep(attempt * 1.5)

    def delete_storage_prefix(self, bucket: str, prefix: str) -> None:
        # Supabase Storage removes exact object paths. The table refresh no longer references
        # older objects, so this best-effort prefix cleanup is allowed to be a no-op.
        try:
            objects = self.list_storage_objects(bucket, prefix)
        except Exception:
            return
        names = ["/".join(part for part in [prefix, item.get("name", "")] if part) for item in objects if item.get("name")]
        if not names:
            return
        self.request(
            "DELETE",
            f"{self.storage_url}/object/{urllib.parse.quote(bucket)}",
            payload={"prefixes": names},
        )

    def list_storage_objects(self, bucket: str, prefix: str) -> list[dict]:
        body = {
            "limit": 1000,
            "offset": 0,
            "prefix": prefix,
        }
        response = self.request(
            "POST",
            f"{self.storage_url}/object/list/{urllib.parse.quote(bucket)}",
            payload=body,
        )
        return json.loads(response.decode("utf-8")) if response else []

    def request(
        self,
        method: str,
        url: str,
        payload=None,
        body: bytes | None = None,
        headers: dict[str, str] | None = None,
    ) -> bytes:
        request_headers = {
            "apikey": self.key,
            "Authorization": f"Bearer {self.key}",
        }
        if payload is not None:
            body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
            request_headers["Content-Type"] = "application/json"
        request_headers.update(headers or {})

        request = urllib.request.Request(url, data=body, method=method, headers=request_headers)
        try:
            with urllib.request.urlopen(request, timeout=60) as response:
                return response.read()
        except urllib.error.HTTPError as exc:
            detail = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Supabase request failed ({method} {url}): {exc.code} {detail}") from exc


def supabase_row(
    record: ProductRecord,
    run_id: str,
    bucket: str,
    storage_path: str,
    public_storage_url: str,
) -> dict:
    return {
        "source_system": SOURCE_SYSTEM,
        "run_id": run_id,
        "catalog_date": record.catalog_date,
        "source_pdf": record.source_pdf,
        "source_pdf_path": record.source_pdf_path,
        "page": record.page,
        "card_index": record.card_index,
        "bbox": record.bbox,
        "model_code": record.model_code or None,
        "product_code": record.product_code or None,
        "material_type": record.material_type or None,
        "price": parse_decimal(record.price),
        "currency": record.currency or None,
        "extraction_status": record.extraction_status,
        "native_text": record.native_text,
        "ocr_text": record.ocr_text,
        "combined_text": record.combined_text,
        "image_sha256": record.image_sha256,
        "image_fingerprint": record.image_fingerprint,
        "image_storage_bucket": bucket,
        "image_storage_path": storage_path,
        "image_url": f"{public_storage_url}/{urllib.parse.quote(bucket)}/{quote_storage_path(storage_path)}",
        "is_visible": bool(record.is_visible),
        "embedding_model": record.embedding_model,
        "embedding_path": record.embedding_path,
        "embedding_preview": record.embedding_preview,
    }


def parse_decimal(value: str) -> float | None:
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def normalize_storage_prefix(value: str) -> str:
    return value.strip().strip("/")


def quote_storage_path(value: str) -> str:
    return "/".join(urllib.parse.quote(part) for part in value.strip("/").split("/"))


def write_excel(records: list[ProductRecord], workbook_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as ExcelImage
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = [
        "catalog_date",
        "source_pdf",
        "page",
        "card_index",
        "model_code",
        "product_code",
        "material_type",
        "price",
        "currency",
        "is_visible",
        "extraction_status",
        "product_image",
        "image_path",
        "image_fingerprint",
        "image_sha256",
        "embedding_model",
        "embedding_path",
        "embedding_preview",
        "bbox",
        "combined_text",
        "native_text",
        "ocr_text",
        "source_pdf_path",
    ]
    ws.append(headers)
    for record in records:
        row = asdict(record)
        ws.append([excel_cell_value(row.get(header, "")) if header != "product_image" else "" for header in headers])

    header_fill = PatternFill("solid", fgColor="1A1510")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    widths = {
        "A": 14,
        "B": 34,
        "C": 8,
        "D": 10,
        "E": 16,
        "F": 18,
        "G": 14,
        "H": 10,
        "I": 14,
        "J": 24,
        "K": 16,
        "L": 52,
        "M": 24,
        "N": 24,
        "O": 20,
        "P": 52,
        "Q": 44,
        "R": 18,
        "S": 70,
        "T": 45,
        "U": 45,
        "V": 52,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    add_product_images(
        ws,
        records,
        ExcelImage,
        thumbnail_dir=workbook_path.parent / "thumbnails" / "latest",
        image_column="L",
        start_row=2,
    )
    ws.column_dimensions["M"].hidden = True
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if records:
        table_range = f"A1:{get_column_letter(len(headers))}{len(records) + 1}"
        table = Table(displayName="ProductsTable", ref=table_range)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    summary = wb.create_sheet("Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Catalog date", records[0].catalog_date if records else ""])
    summary.append(["Product card candidates", len(records)])
    summary.append(["Source PDFs", len({record.source_pdf for record in records})])
    summary.append(["With product code", sum(1 for record in records if record.product_code)])
    summary.append(["With price", sum(1 for record in records if record.price)])
    summary.append(["Need OCR for prefix", sum(1 for record in records if "needs_ocr" in record.extraction_status)])
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 24
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.freeze_panes = "A2"

    wb.save(workbook_path)


def add_product_images(
    ws,
    records: list[ProductRecord],
    excel_image_cls,
    thumbnail_dir: Path,
    image_column: str,
    start_row: int,
) -> None:
    thumbnail_dir.mkdir(parents=True, exist_ok=True)

    for row_index, record in enumerate(records, start=start_row):
        image_path = Path(record.image_path)
        if not image_path.exists():
            continue

        thumbnail_path = create_excel_thumbnail(image_path, thumbnail_dir)
        image = excel_image_cls(str(thumbnail_path))
        width, height = thumbnail_dimensions(thumbnail_path)
        image.width = width
        image.height = height
        ws.row_dimensions[row_index].height = max(72, height * 0.75 + 8)
        ws.add_image(image, f"{image_column}{row_index}")


def create_excel_thumbnail(
    image_path: Path,
    thumbnail_dir: Path,
    max_width: int = 96,
    max_height: int = 96,
) -> Path:
    from PIL import Image

    with Image.open(image_path) as image:
        image = image.convert("RGB")
        image.thumbnail((max_width, max_height))
        thumbnail_path = thumbnail_dir / f"{safe_filename(image_path.stem)}.jpg"
        image.save(thumbnail_path, "JPEG", quality=72, optimize=True)
        return thumbnail_path


def create_supabase_image(
    image_path: Path,
    storage_image_dir: Path,
    max_width: int = 900,
    max_height: int = 1250,
) -> Path:
    from PIL import Image

    storage_image_dir.mkdir(parents=True, exist_ok=True)
    with Image.open(image_path) as image:
        image = image.convert("RGB")
        image.thumbnail((max_width, max_height))
        output_path = storage_image_dir / f"{safe_filename(image_path.stem)}.jpg"
        image.save(output_path, "JPEG", quality=82, optimize=True, progressive=True)
        return output_path


def thumbnail_dimensions(image_path: Path) -> tuple[int, int]:
    from PIL import Image

    with Image.open(image_path) as image:
        return image.size


def write_csv(records: list[ProductRecord], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = list(ProductRecord.__dataclass_fields__.keys())
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(asdict(record))


def excel_cell_value(value):
    if isinstance(value, str):
        return EXCEL_ILLEGAL_CHARACTERS.sub("", value)[:32767]
    return value


if __name__ == "__main__":
    sys.exit(main())
